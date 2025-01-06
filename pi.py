import os
import base64
import openpyxl
import string

def base64_to_png(base64_string, file_name, save_folder):
    # Verificar se a string Base64 está vazia
    if not base64_string:
        print(f"String Base64 vazia para '{file_name}'. Ignorando esta entrada.")
        return

    # Remover caracteres de espaço em branco da string Base64
    base64_string = base64_string.replace(" ", "")

    # Decodificar a string Base64 para bytes
    try:
        image_data = base64.b64decode(base64_string)
    except base64.binascii.Error:
        # Adicionar o padding à string Base64 e tentar decodificar novamente
        padding_needed = 4 - len(base64_string) % 4
        base64_string += '=' * padding_needed
        try:
            image_data = base64.b64decode(base64_string)
        except Exception as e:
            print(f"Erro ao decodificar Base64 para '{file_name}': {e}")
            return

    # Substituir caracteres inválidos em nomes de arquivo
    valid_chars = "-_() %s%s" % (string.ascii_letters, string.digits)
    file_name = ''.join(c if c in valid_chars else '-' for c in file_name)

    # Caminho do arquivo PNG na pasta de salvamento
    file_path = os.path.join(save_folder, f"{file_name}.png")

    # Verificar se o arquivo já existe
    if os.path.exists(file_path):
        print(f"O arquivo '{file_path}' já existe. Ignorando esta entrada.")
        return

    # Salvar os bytes como um arquivo PNG
    with open(file_path, "wb") as f:
        f.write(image_data)

    return file_path

def main():
    # Caminho do arquivo Excel
    excel_file = r"C:\site\dada.xlsx"

    # Pasta de salvamento
    save_folder = os.path.join(os.path.dirname(excel_file), "ft")

    # Cria a pasta de salvamento se não existir
    if not os.path.exists(save_folder):
        os.makedirs(save_folder)

    # Carrega o arquivo Excel e seleciona a aba 'DADA'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["DADA"]

    # Adiciona cabeçalhos para as novas colunas
    ws["N1"] = "Caminho do Arquivo"

    # Percorre as linhas do arquivo Excel
    row_number = 2  # Inicia em 2 para pular a linha do cabeçalho
    for row in ws.iter_rows(min_row=2, values_only=True):
        base64_string, file_name = row[2], row[11]
        file_path = base64_to_png(base64_string, file_name, save_folder)
        if file_path:
            ws.cell(row=row_number, column=14, value=file_path)
        row_number += 1

    # Salva as alterações no arquivo Excel
    wb.save(excel_file)

    print("Todas as imagens foram convertidas e os caminhos foram salvos com sucesso!")

if __name__ == "__main__":
    main()
