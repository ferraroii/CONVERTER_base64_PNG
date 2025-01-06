import os
import base64
import openpyxl
from unidecode import unidecode
import re
import string


def normalize_filename(name):
    """
    Remove acentos, caracteres especiais e formata o nome para uso como nome de arquivo.
    """
    if not name:
        return ""
    name = unidecode(name)  # Remove acentos
    name = re.sub(r'\s+', '_', name)  # Substitui espaços por underscores
    name = re.sub(r'[^\w\s.-]', '', name)  # Remove caracteres inválidos
    return name


def base64_to_png(base64_string, file_name, save_folder):
    """
    Converte uma string Base64 em uma imagem PNG e a salva na pasta especificada.
    """
    if not base64_string:
        print(f"String Base64 vazia para '{file_name}'. Ignorando esta entrada.")
        return None

    # Corrige a string Base64
    base64_string = base64_string.replace(" ", "")
    try:
        image_data = base64.b64decode(base64_string)
    except base64.binascii.Error:
        padding_needed = 4 - len(base64_string) % 4
        base64_string += '=' * padding_needed
        try:
            image_data = base64.b64decode(base64_string)
        except Exception as e:
            print(f"Erro ao decodificar Base64 para '{file_name}': {e}")
            return None

    # Substituir caracteres inválidos em nomes de arquivo
    valid_chars = "-_() %s%s" % (string.ascii_letters, string.digits)
    file_name = ''.join(c if c in valid_chars else '-' for c in file_name)

    # Caminho completo para salvar o arquivo
    file_path = os.path.join(save_folder, f"{file_name}.png")

    # Verificar se o arquivo já existe
    if os.path.exists(file_path):
        print(f"O arquivo '{file_path}' já existe. Ignorando esta entrada.")
        return None

    # Salvar a imagem como arquivo PNG
    with open(file_path, "wb") as f:
        f.write(image_data)

    return file_path


def main():
    # Caminho do arquivo Excel
    excel_file = r"C:\site\dada.xlsx"

    # Pasta de salvamento das imagens
    save_folder = os.path.join(os.path.dirname(excel_file), "ft")

    # Cria a pasta se não existir
    if not os.path.exists(save_folder):
        os.makedirs(save_folder)

    # Carrega o arquivo Excel
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["DADA"]

    # Adiciona cabeçalhos para as novas colunas
    ws["C1"] = "Nome Normalizado"
    ws["N1"] = "Caminho do Arquivo"

    # Processa as linhas da planilha
    for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        col_a_value = row[0].value  # Coluna A
        col_b_value = row[2].value  # Base64 na coluna C
        col_c_value = row[11].value  # Nome na coluna L

        # Normaliza o valor da coluna A
        if isinstance(col_a_value, str):
            normalized_name = normalize_filename(col_a_value)
            ws[f"C{row_number}"].value = normalized_name

        # Converte Base64 em PNG
        if col_b_value and col_c_value:
            file_path = base64_to_png(col_b_value, col_c_value, save_folder)
            if file_path:
                ws[f"N{row_number}"].value = file_path

    # Salva as alterações no arquivo Excel
    wb.save(excel_file)
    print("Processamento concluído com sucesso!")


if __name__ == "__main__":
    main()
